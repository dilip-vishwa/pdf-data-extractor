from excel_utils.excel_style_utils.excel_color_utils import check_if_given_cell_is_text_or_green_or_empty
from openpyxl.cell.cell import Cell

label_already_parsed = []

def get_cell_far_in_given_direction_of_the_given_cell(cell, direction='left', distance=2, sheet=None, max_parsing_times=None) -> Cell|None:
    if isinstance(cell, str):
        cell = sheet[cell]
    
    left_cell = None
    top_cell = None
    right_cell = None
    bottom_cell = None
    if direction == 'left' or direction == 'all':
        if cell.column - distance < 1:
            left_cell = None
        else:
            left_cell = sheet.cell(row=cell.row, column=cell.column - distance)
            type_of_cell = check_if_given_cell_is_text_or_green_or_empty(left_cell, sheet)
            if type_of_cell == 'empty':
                if max_parsing_times['left'] == 0:
                    left_cell = None
                else:
                    max_parsing_times['left'] -= 1
                    left_cell = get_cell_far_in_given_direction_of_the_given_cell(left_cell, direction='left', distance=distance, sheet=sheet, max_parsing_times=max_parsing_times)
        
    if direction == 'top' or direction == 'all':
        if cell.row - distance < 1:
            top_cell = None
        else:
            top_cell = sheet.cell(row=cell.row - distance, column=cell.column)
            type_of_cell = check_if_given_cell_is_text_or_green_or_empty(top_cell, sheet)
            if type_of_cell == 'empty':
                if max_parsing_times['top'] == 0:
                    top_cell = None
                else:
                    max_parsing_times['top'] -= 1
                    top_cell = get_cell_far_in_given_direction_of_the_given_cell(top_cell, direction='top', distance=distance, sheet=sheet, max_parsing_times=max_parsing_times)
        
    if direction == 'right' or direction == 'all':
        if cell.column + distance > sheet.max_column:
            right_cell = None
        else:
            right_cell = sheet.cell(row=cell.row, column=cell.column + distance)
            type_of_cell = check_if_given_cell_is_text_or_green_or_empty(right_cell, sheet)
            if type_of_cell == 'empty':
                if max_parsing_times['right'] == 0:
                    right_cell = None
                else:
                    max_parsing_times['right'] -= 1
                    right_cell = get_cell_far_in_given_direction_of_the_given_cell(right_cell, direction='right', distance=distance, sheet=sheet, max_parsing_times=max_parsing_times)
    
    if direction == 'bottom' or direction == 'all':
        if cell.row + distance > sheet.max_row:
            bottom_cell = None
        else:
            bottom_cell = sheet.cell(row=cell.row + distance, column=cell.column)
            type_of_cell = check_if_given_cell_is_text_or_green_or_empty(bottom_cell, sheet)
            if type_of_cell == 'empty':
                if max_parsing_times['bottom'] == 0:
                    bottom_cell = None
                else:
                    max_parsing_times['bottom'] -= 1
                    bottom_cell = get_cell_far_in_given_direction_of_the_given_cell(bottom_cell, direction='bottom', distance=distance, sheet=sheet, max_parsing_times=max_parsing_times)
        
    if direction == 'left':
        return left_cell
    elif direction == 'top':
        return top_cell
    elif direction == 'right':
        return right_cell
    elif direction == 'bottom':
        return bottom_cell
    if direction == 'all':
        return (left_cell,
            top_cell,
            right_cell,
            bottom_cell)
