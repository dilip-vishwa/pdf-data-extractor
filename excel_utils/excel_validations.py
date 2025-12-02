import openpyxl
import json
import sys
from openpyxl.utils import range_boundaries, get_column_letter

def get_validation_map(wb, ws):
    """
    Creates a map of cell coordinate -> list of validation options.
    Resolves references to other sheets.
    """
    validation_map = {}
    
    if not ws.data_validations:
        return validation_map

    for val in ws.data_validations.dataValidation:
        if val.type == 'list':
            formula = val.formula1
            options = []
            
            # Resolve formula
            if formula:
                # Case 1: Literal list "Option1,Option2"
                if ',' in formula and not formula.startswith('='):
                    options = [x.strip() for x in formula.split(',')]
                
                # Case 2: Reference "=Sheet!A1:A10" or "=$A$1:$A$10"
                else:
                    try:
                        target_sheet = ws
                        ref = formula
                        
                        # Handle sheet reference
                        if '!' in formula:
                            sheet_part, ref = formula.split('!')
                            sheet_part = sheet_part.replace("'", "")
                            if sheet_part in wb.sheetnames:
                                target_sheet = wb[sheet_part]
                        
                        # Clean reference
                        ref = ref.replace('$', '').replace('=', '')
                        
                        # Fetch values
                        if ':' in ref:
                            rows = target_sheet[ref]
                            # Normalize to tuple of tuples
                            if not isinstance(rows, tuple):
                                rows = ((rows,),)
                            elif not isinstance(rows[0], tuple):
                                rows = (rows,)
                                
                            for row in rows:
                                for cell in row:
                                    if cell.value is not None:
                                        options.append(str(cell.value))
                        else:
                            # Single cell
                            cell = target_sheet[ref]
                            if cell.value is not None:
                                options.append(str(cell.value))
                                
                    except Exception as e:
                        # print(f"Warning: Could not resolve validation {formula}: {e}", file=sys.stderr)
                        pass

            # Expand range (e.g. "A1:A5 A7") to individual cells
            # sqref can contain multiple ranges separated by space
            ranges = str(val.sqref).split()
            for rng in ranges:
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(rng)
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            coord = f"{get_column_letter(col)}{row}"
                            validation_map[coord] = options
                except ValueError:
                    # Handle single cell ref if range_boundaries fails (though it shouldn't for valid refs)
                    validation_map[str(val.sqref)] = options

    return validation_map

if __name__ == "__main__":
    file_path = "/home/adming/project/pdf-data-extractor/dataset/SPECIFICATION_SUMMARY_SHEET-BLANK.xlsx"
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    validation_map = get_validation_map(wb, ws)
    print(validation_map)
