from openpyxl import load_workbook
# from openpyxl.worksheet.datavalidation import DataValidation

def read_excel_file_and_create_a_pandas_dataframe(excel_file_path):
    with open(excel_file_path, 'rb') as f:
        df = pd.read_excel(f)
    return df

# def read_excel_file_using_openpyxl(excel_file_path):
#     wbs = load_workbook(excel_file_path)
#     for wb in wbs:
#         wb # merged_cells, dimensions, title
#     ws = wb["Sheet"]
#     for 
#     for row in ws.iter_rows(values_only=True):
#         print(row)


def read_data_validation_from_excel(excel_file_path):
    # Load the workbook
    file_path = excel_file_path
    wb = load_workbook(file_path, data_only=False)
    ws = wb['R0']  # or wb['SheetName']

    # --- Read cell value (selected value in dropdown) ---
    cell = ws["E21"]   # Example: cell with dropdown
    print("Selected value in dropdown:", cell.value)

    # --- Read the dropdown list items from data validation ---
    dropdown_items = []

    for dv in ws.data_validations.dataValidation:
        # Check if the cell is inside the data validation range
        for r in dv.ranges:
            # if "E21" in r:   # you can test any cell address
                # The formula1 contains the dropdown source
            formula = dv.formula1
            print("Dropdown formula:", formula)

            # Case 1: dropdown uses a literal list "A,B,C"
            if '"' in formula:
                dropdown_items = formula.replace('"', '').split(',')
            
            # Case 2: dropdown refers to a cell range: e.g. =$D$1:$D$5
            elif "!" in formula or ":" in formula:
                # Remove sheet name if present
                if "!" in formula:
                    sheet_name, cell_range = formula.split("!")
                    sheet_name = sheet_name.replace("=", "")
                    sheet = wb[sheet_name]
                else:
                    sheet = ws
                    cell_range = formula.replace("=", "")

                # Load items from range
                for row in sheet[cell_range]:
                    for c in row:
                        dropdown_items.append(c.value)

    print("Dropdown items:", dropdown_items)


def read_validation_data_new(excel_file_path):
    file_path = excel_file_path
    wb = load_workbook(file_path, data_only=False)  # IMPORTANT: data_only=False
    ws = wb.active

    cell_address = "D17"   # Your dropdown cell

    # ---------------------------------------------
    # 1. Read the selected value in the dropdown
    # ---------------------------------------------
    selected_value = ws[cell_address].value
    print("Selected dropdown value:", selected_value)

    # ---------------------------------------------
    # 2. Read dropdown LIST items from data validation
    # ---------------------------------------------
    dropdown_list = []

    for dv in ws.data_validations.dataValidation:
        # Check if the cell is inside this validation's range
        # if any(cell_address in str(rng) for rng in dv.ranges):
        formula = dv.formula1
        print(f"Data validation formula: {formula} in {[d.coord for d in dv.cells]}")

    #     # Case A: Literal list "Red,Green,Blue"
    #     if formula.startswith('"') and formula.endswith('"'):
    #         dropdown_list = formula.strip('"').split(",")

    #     # Case B: Cell range =Sheet1!$D$1:$D$5
    #     else:
    #         formula = formula.replace("=", "")
    #         if "!" in formula:
    #             sheet_name, cell_range = formula.split("!")
    #             sheet = wb[sheet_name]
    #         else:
    #             sheet = ws
    #             cell_range = formula

    #         for row in sheet[cell_range]:
    #             for c in row:
    #                 dropdown_list.append(c.value)

    # print("Dropdown items:", dropdown_list)


if __name__ == "__main__":
    # excel_file_path = "/home/adming/project/pdf-data-extractor/SPECIFICATION_SUMMARY_SHEET-BLANK.xlsx"
    excel_file_path = "/home/adming/project/pdf-data-extractor/SPECIFICATION_SUMMARY_SHEET-BLANK.xlsx"
    # df = read_excel_file_and_create_a_pandas_dataframe(excel_file_path)
    # print(df)
    # read_excel_file_using_openpyxl(excel_file_path)
    # read_data_validation_from_excel(excel_file_path)
    read_validation_data_new(excel_file_path)